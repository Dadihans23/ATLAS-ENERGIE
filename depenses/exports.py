"""
Exports CSV et Excel pour les dépenses.
- CSV  : séparateur point-virgule, encodage UTF-8 BOM (compatible Excel FR)
- Excel: fichier .xlsx formaté avec openpyxl (couleurs, totaux, colonnes auto)

Sécurité (HIGH-05) : exports limités à MAX_EXPORT_ROWS lignes pour éviter
les dénis de service par export massif.
"""
import csv
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone

from .models import DepenseExploitation, DepenseFraisGeneraux

MAX_EXPORT_ROWS = 10_000  # Limite de sécurité anti-DoS


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _qs_exploitation(user):
    qs = DepenseExploitation.objects.select_related(
        'centre_budgetaire', 'created_by'
    ).order_by('date_saisie')
    if not user.is_chef:
        qs = qs.filter(created_by=user)
    return qs


def _qs_frais_generaux(user):
    qs = DepenseFraisGeneraux.objects.select_related(
        'created_by'
    ).order_by('date_saisie')
    if not user.is_chef:
        qs = qs.filter(created_by=user)
    return qs


def _format_xof(val):
    """Retourne un float pour Excel (sans unité)."""
    try:
        return float(Decimal(str(val)))
    except Exception:
        return 0.0


def _check_export_limit(request, qs, redirect_url):
    """
    Vérifie que le queryset ne dépasse pas MAX_EXPORT_ROWS.
    Retourne True si la limite est dépassée (et affiche un message d'erreur).
    """
    count = qs.count()
    if count > MAX_EXPORT_ROWS:
        messages.error(
            request,
            f"Export limité à {MAX_EXPORT_ROWS:,} lignes. "
            f"Votre sélection contient {count:,} enregistrements — filtrez d'abord."
        )
        return True
    return False


# ─── CSV ──────────────────────────────────────────────────────────────────────

@login_required
def export_exploitation_csv(request):
    qs = _qs_exploitation(request.user)
    if _check_export_limit(request, qs, 'depenses:exploitation_liste'):
        return redirect(reverse('depenses:exploitation_liste'))

    today = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig',
    )
    response['Content-Disposition'] = f'attachment; filename="depenses_exploitation_{today}.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'N° Saisie', 'Date', 'Projet', 'Ligne budgétaire', 'Nature',
        'Libellé', 'Montant', 'Devise', 'Montant XOF',
        'Taux change', 'Fournisseur', 'N° Pièce', 'Commentaire', 'Saisi par',
    ])

    total_xof = Decimal('0.00')
    for d in qs:
        total_xof += d.montant_xof
        writer.writerow([
            d.numero_saisie,
            d.date_saisie.strftime('%d/%m/%Y'),
            d.centre_budgetaire.nom,
            d.get_ligne_budgetaire_display(),
            d.get_nature_depense_display(),
            d.libelle,
            str(d.montant).replace('.', ','),
            d.devise,
            str(d.montant_xof).replace('.', ','),
            str(d.taux_change_utilise).replace('.', ','),
            d.fournisseur,
            d.numero_piece,
            d.commentaire,
            d.created_by.get_full_name(),
        ])

    writer.writerow([])
    writer.writerow(['', '', '', '', '', 'TOTAL XOF', '', '', str(total_xof).replace('.', ',')])
    return response


@login_required
def export_frais_generaux_csv(request):
    qs = _qs_frais_generaux(request.user)
    if _check_export_limit(request, qs, 'depenses:fg_liste'):
        return redirect(reverse('depenses:fg_liste'))

    today = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig',
    )
    response['Content-Disposition'] = f'attachment; filename="frais_generaux_{today}.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'N° Saisie', 'Date', 'Ligne budgétaire', 'Nature',
        'Libellé', 'Montant', 'Devise', 'Montant XOF',
        'Fournisseur', 'N° Pièce', 'Commentaire', 'Saisi par',
    ])

    total_xof = Decimal('0.00')
    for d in qs:
        total_xof += d.montant_xof
        writer.writerow([
            d.numero_saisie,
            d.date_saisie.strftime('%d/%m/%Y'),
            d.get_ligne_budgetaire_display(),
            d.get_nature_depense_display(),
            d.libelle,
            str(d.montant).replace('.', ','),
            d.devise,
            str(d.montant_xof).replace('.', ','),
            d.fournisseur,
            d.numero_piece,
            d.commentaire,
            d.created_by.get_full_name(),
        ])

    writer.writerow([])
    writer.writerow(['', '', '', '', 'TOTAL XOF', '', '', str(total_xof).replace('.', ',')])
    return response


# ─── Excel ────────────────────────────────────────────────────────────────────

@login_required
def export_exploitation_excel(request):
    qs = _qs_exploitation(request.user)
    if _check_export_limit(request, qs, 'depenses:exploitation_liste'):
        return redirect(reverse('depenses:exploitation_liste'))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installé.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dépenses Exploitation"

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB'),
    )
    align_right = Alignment(horizontal='right')
    align_center = Alignment(horizontal='center')

    headers = [
        'N° Saisie', 'Date', 'Projet', 'Ligne budgétaire', 'Nature',
        'Libellé', 'Montant', 'Devise', 'Montant XOF', 'Fournisseur',
        'N° Pièce', 'Saisi par',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    row_num = 2
    total_xof = Decimal('0.00')
    for d in qs:
        total_xof += d.montant_xof
        row_data = [
            d.numero_saisie,
            d.date_saisie,
            d.centre_budgetaire.nom,
            d.get_ligne_budgetaire_display(),
            d.get_nature_depense_display(),
            d.libelle,
            _format_xof(d.montant),
            d.devise,
            _format_xof(d.montant_xof),
            d.fournisseur,
            d.numero_piece,
            d.created_by.get_full_name(),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in (7, 9):
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
        row_num += 1

    ws.cell(row=row_num, column=8, value='TOTAL XOF').font = total_font
    total_cell = ws.cell(row=row_num, column=9, value=_format_xof(total_xof))
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = align_right
    total_cell.number_format = '#,##0.00'

    col_widths = [16, 12, 28, 22, 20, 40, 14, 8, 16, 22, 14, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="depenses_exploitation_{today}.xlsx"'
    return response


@login_required
def export_frais_generaux_excel(request):
    qs = _qs_frais_generaux(request.user)
    if _check_export_limit(request, qs, 'depenses:fg_liste'):
        return redirect(reverse('depenses:fg_liste'))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installé.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Frais Généraux"

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4C1D95', end_color='4C1D95', fill_type='solid')
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='E5E7EB'))
    align_right = Alignment(horizontal='right')
    align_center = Alignment(horizontal='center')

    headers = [
        'N° Saisie', 'Date', 'Ligne budgétaire', 'Nature',
        'Libellé', 'Montant', 'Devise', 'Montant XOF',
        'Fournisseur', 'N° Pièce', 'Saisi par',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    row_num = 2
    total_xof = Decimal('0.00')
    for d in qs:
        total_xof += d.montant_xof
        row_data = [
            d.numero_saisie,
            d.date_saisie,
            d.get_ligne_budgetaire_display(),
            d.get_nature_depense_display(),
            d.libelle,
            _format_xof(d.montant),
            d.devise,
            _format_xof(d.montant_xof),
            d.fournisseur,
            d.numero_piece,
            d.created_by.get_full_name(),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in (6, 8):
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
        row_num += 1

    ws.cell(row=row_num, column=7, value='TOTAL XOF').font = total_font
    total_cell = ws.cell(row=row_num, column=8, value=_format_xof(total_xof))
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.alignment = align_right
    total_cell.number_format = '#,##0.00'

    col_widths = [16, 12, 28, 20, 40, 14, 8, 16, 22, 14, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="frais_generaux_{today}.xlsx"'
    return response
